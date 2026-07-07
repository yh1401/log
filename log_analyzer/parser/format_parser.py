"""统一的日志格式解析器 - 支持多种文件格式."""

import os
import re
import csv
import json
import xml.etree.ElementTree as ET
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime
from typing import List, Dict, Any, Optional, Iterator, AsyncIterator, Type
from enum import Enum
import logging

logger = logging.getLogger(__name__)


class FormatType(Enum):
    """支持的日志格式类型"""
    LOG = "log"           # 标准日志格式
    JSON = "json"         # JSON格式
    CSV = "csv"           # CSV格式
    XML = "xml"           # XML格式
    EXCEL = "excel"       # Excel格式
    PDF = "pdf"           # PDF格式
    TXT = "txt"           # 纯文本格式
    UNKNOWN = "unknown"   # 未知格式


@dataclass
class LogEntry:
    """标准化日志条目"""
    timestamp: Optional[str] = None
    level: str = "INFO"
    message: str = ""
    class_name: str = ""
    method: str = ""
    device_id: str = ""
    trace_id: str = ""
    raw_data: Dict[str, Any] = field(default_factory=dict)
    line_number: int = 0
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "timestamp": self.timestamp,
            "level": self.level,
            "message": self.message,
            "class": self.class_name,
            "method": self.method,
            "device_id": self.device_id,
            "trace_id": self.trace_id,
            "line_number": self.line_number,
            **self.raw_data
        }


class FormatParser(ABC):
    """格式解析器抽象基类"""
    
    format_type: FormatType = FormatType.UNKNOWN
    extensions: List[str] = []
    
    @abstractmethod
    def parse(self, file_path: str, chunk_size: int = 10000) -> Iterator[List[LogEntry]]:
        """同步解析文件，返回条目块迭代器"""
        pass
    
    @abstractmethod
    async def parse_async(self, file_path: str, chunk_size: int = 10000) -> AsyncIterator[List[LogEntry]]:
        """异步解析文件，返回条目块异步迭代器"""
        pass
    
    @abstractmethod
    def detect_format(self, file_path: str) -> bool:
        """检测文件是否符合此解析器的格式"""
        pass


class LogFormatParser(FormatParser):
    """标准日志格式解析器 - 兼容原有解析器"""
    
    format_type = FormatType.LOG
    extensions = [".log", ".txt"]
    
    LOG_PATTERN = re.compile(
        r'^(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}(?:\.\d{3})?)'
        r'\s+\[([^\]]+)\]'
        r'\s+(\w+)'
        r'\s+([a-f0-9-]+)'
        r'\s+([^\s-]+)'
        r'\s+-\s+(.*)$'
    )
    
    JAVA_LOG_PATTERN = re.compile(
        r'^\[(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2},\d{3})\]'
        r'\s+\[(\w+)\s*\]'
        r'\s+\[LOGID:([^\]]+)\]'
        r'\s+([^\s]+)'
        r'\s+([^\s]+)'
        r'\s+-\s+(.*)$'
    )
    
    SIMPLE_LOG_PATTERN = re.compile(
        r'^(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}(?:\.\d{3})?)'
        r'\s+(\w+)'
        r'\s+(.*)$'
    )
    
    def detect_format(self, file_path: str) -> bool:
        ext = os.path.splitext(file_path)[1].lower()
        if ext not in self.extensions:
            return False
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                first_lines = [f.readline() for _ in range(10)]
                return any(self._try_parse_line(line) for line in first_lines if line.strip())
        except Exception:
            return False
    
    def _try_parse_line(self, line: str) -> Optional[LogEntry]:
        """尝试解析单行日志"""
        for pattern in [self.LOG_PATTERN, self.JAVA_LOG_PATTERN, self.SIMPLE_LOG_PATTERN]:
            match = pattern.match(line.strip())
            if match:
                groups = match.groups()
                if len(groups) == 6:
                    return LogEntry(
                        timestamp=groups[0],
                        level=groups[2],
                        message=groups[5],
                        class_name=groups[4],
                        trace_id=groups[3],
                        line_number=0
                    )
                elif len(groups) == 3:
                    return LogEntry(
                        timestamp=groups[0],
                        level=groups[1],
                        message=groups[2],
                        line_number=0
                    )
        return None
    
    def parse(self, file_path: str, chunk_size: int = 10000) -> Iterator[List[LogEntry]]:
        chunk: List[LogEntry] = []
        line_num = 0
        
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                line_num += 1
                entry = self._try_parse_line(line)
                if entry:
                    entry.line_number = line_num
                    chunk.append(entry)
                
                if len(chunk) >= chunk_size:
                    yield chunk
                    chunk = []
        
        if chunk:
            yield chunk
    
    async def parse_async(self, file_path: str, chunk_size: int = 10000) -> AsyncIterator[List[LogEntry]]:
        import aiofiles
        chunk: List[LogEntry] = []
        line_num = 0
        
        async with aiofiles.open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            async for line in f:
                line_num += 1
                entry = self._try_parse_line(line)
                if entry:
                    entry.line_number = line_num
                    chunk.append(entry)
                
                if len(chunk) >= chunk_size:
                    yield chunk
                    chunk = []
        
        if chunk:
            yield chunk


class JSONFormatParser(FormatParser):
    """JSON格式日志解析器"""
    
    format_type = FormatType.JSON
    extensions = [".json"]
    
    def detect_format(self, file_path: str) -> bool:
        ext = os.path.splitext(file_path)[1].lower()
        if ext != ".json":
            return False
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read(1000)
                json.loads(content)
                return True
        except Exception:
            return False
    
    def _parse_json_entry(self, data: Dict[str, Any], line_num: int) -> LogEntry:
        """解析JSON对象为LogEntry"""
        # 尝试多种时间字段名
        timestamp = (
            data.get("timestamp") or 
            data.get("time") or 
            data.get("@timestamp") or
            data.get("datetime") or
            data.get("date")
        )
        
        # 尝试多种级别字段名
        level = (
            data.get("level") or 
            data.get("severity") or 
            data.get("log_level") or
            data.get("loglevel") or
            "INFO"
        ).upper()
        
        # 尝试多种消息字段名
        message = (
            data.get("message") or 
            data.get("msg") or 
            data.get("content") or
            data.get("text") or
            str(data)
        )
        
        # 尝试多种类名字段名
        class_name = (
            data.get("class") or 
            data.get("class_name") or 
            data.get("logger") or
            data.get("source")
        ) or ""
        
        # 尝试多种方法名字段名
        method = (
            data.get("method") or 
            data.get("function") or
            data.get("func")
        ) or ""
        
        return LogEntry(
            timestamp=str(timestamp) if timestamp else None,
            level=level,
            message=str(message),
            class_name=class_name,
            method=method,
            raw_data=data,
            line_number=line_num
        )
    
    def parse(self, file_path: str, chunk_size: int = 10000) -> Iterator[List[LogEntry]]:
        chunk: List[LogEntry] = []
        line_num = 0
        
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            line_num = 0
            
            # 尝试解析为JSON数组
            try:
                data = json.loads(content)
                if isinstance(data, list):
                    for item in data:
                        line_num += 1
                        entry = self._parse_json_entry(item, line_num)
                        chunk.append(entry)
                        if len(chunk) >= chunk_size:
                            yield chunk
                            chunk = []
                elif isinstance(data, dict):
                    line_num = 1
                    entry = self._parse_json_entry(data, line_num)
                    chunk.append(entry)
                    if len(chunk) >= chunk_size:
                        yield chunk
                        chunk = []
            except json.JSONDecodeError:
                # 尝试JSONL格式（每行一个JSON）
                for line in content.split('\n'):
                    line_num += 1
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        item = json.loads(line)
                        entry = self._parse_json_entry(item, line_num)
                        chunk.append(entry)
                        if len(chunk) >= chunk_size:
                            yield chunk
                            chunk = []
                    except json.JSONDecodeError:
                        continue
        
        if chunk:
            yield chunk
    
    async def parse_async(self, file_path: str, chunk_size: int = 10000) -> AsyncIterator[List[LogEntry]]:
        import aiofiles
        chunk: List[LogEntry] = []
        line_num = 0
        
        async with aiofiles.open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = await f.read()
            line_num = 0
            
            try:
                data = json.loads(content)
                if isinstance(data, list):
                    for item in data:
                        line_num += 1
                        entry = self._parse_json_entry(item, line_num)
                        chunk.append(entry)
                        if len(chunk) >= chunk_size:
                            yield chunk
                            chunk = []
                elif isinstance(data, dict):
                    line_num = 1
                    entry = self._parse_json_entry(data, line_num)
                    chunk.append(entry)
                    if chunk:
                        yield chunk
            except json.JSONDecodeError:
                for line in content.split('\n'):
                    line_num += 1
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        item = json.loads(line)
                        entry = self._parse_json_entry(item, line_num)
                        chunk.append(entry)
                        if len(chunk) >= chunk_size:
                            yield chunk
                            chunk = []
                    except json.JSONDecodeError:
                        continue
        
        if chunk:
            yield chunk


class CSVFormatParser(FormatParser):
    """CSV格式日志解析器"""
    
    format_type = FormatType.CSV
    extensions = [".csv"]
    
    def detect_format(self, file_path: str) -> bool:
        ext = os.path.splitext(file_path)[1].lower()
        if ext != ".csv":
            return False
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                sample = f.read(500)
                # 尝试解析为CSV
                lines = sample.strip().split('\n')
                if len(lines) >= 2:
                    csv.Sniffer().sniff(sample)
                    return True
        except Exception:
            return False
        return False
    
    def _parse_csv_row(self, row: Dict[str, str], line_num: int) -> LogEntry:
        """解析CSV行为LogEntry"""
        # 尝试多种字段映射
        timestamp = (
            row.get("timestamp") or 
            row.get("time") or 
            row.get("datetime") or
            row.get("date") or
            row.get("@timestamp")
        )
        
        level = (
            row.get("level") or 
            row.get("severity") or 
            row.get("log_level")
        ) or "INFO"
        
        message = (
            row.get("message") or 
            row.get("msg") or 
            row.get("content")
        ) or str(row)
        
        class_name = row.get("class") or row.get("logger") or ""
        method = row.get("method") or row.get("function") or ""
        device_id = row.get("device_id") or row.get("deviceId") or ""
        trace_id = row.get("trace_id") or row.get("traceId") or ""
        
        return LogEntry(
            timestamp=timestamp,
            level=level.upper() if level else "INFO",
            message=message,
            class_name=class_name,
            method=method,
            device_id=device_id,
            trace_id=trace_id,
            raw_data=row,
            line_number=line_num
        )
    
    def parse(self, file_path: str, chunk_size: int = 10000) -> Iterator[List[LogEntry]]:
        chunk: List[LogEntry] = []
        line_num = 0
        
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            # 自动检测分隔符
            sample = f.read(8192)
            f.seek(0)
            
            try:
                dialect = csv.Sniffer().sniff(sample)
            except Exception:
                dialect = csv.excel
            
            reader = csv.DictReader(f, dialect=dialect)
            
            for row in reader:
                line_num += 1
                entry = self._parse_csv_row(row, line_num)
                chunk.append(entry)
                
                if len(chunk) >= chunk_size:
                    yield chunk
                    chunk = []
        
        if chunk:
            yield chunk
    
    async def parse_async(self, file_path: str, chunk_size: int = 10000) -> AsyncIterator[List[LogEntry]]:
        import aiofiles
        chunk: List[LogEntry] = []
        line_num = 0
        
        async with aiofiles.open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = await f.read()
        
        import io
        reader = csv.DictReader(io.StringIO(content))
        
        for row in reader:
            line_num += 1
            entry = self._parse_csv_row(row, line_num)
            chunk.append(entry)
            
            if len(chunk) >= chunk_size:
                yield chunk
                chunk = []
        
        if chunk:
            yield chunk


class XMLFormatParser(FormatParser):
    """XML格式日志解析器"""
    
    format_type = FormatType.XML
    extensions = [".xml"]
    
    def detect_format(self, file_path: str) -> bool:
        ext = os.path.splitext(file_path)[1].lower()
        if ext != ".xml":
            return False
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                sample = f.read(500)
                return sample.strip().startswith('<')
        except Exception:
            return False
    
    def _parse_xml_element(self, elem: ET.Element, line_num: int) -> LogEntry:
        """解析XML元素为LogEntry"""
        # 递归收集所有文本内容
        def get_text(e: ET.Element) -> str:
            return e.text or ' '.join(ET.tostring(e, encoding='unicode').split())
        
        data = {child.tag: get_text(child) for child in elem}
        data['_tag'] = elem.tag
        
        # 提取属性
        if elem.attrib:
            data['_attributes'] = dict(elem.attrib)
        
        # 尝试多种字段映射
        timestamp = data.get("timestamp") or data.get("time") or data.get("@timestamp")
        level = (data.get("level") or data.get("severity") or "INFO").upper()
        message = data.get("message") or data.get("msg") or data.get("content") or ET.tostring(elem, encoding='unicode')
        class_name = data.get("class") or data.get("logger") or elem.tag
        
        return LogEntry(
            timestamp=timestamp,
            level=level,
            message=message,
            class_name=class_name,
            raw_data=data,
            line_number=line_num
        )
    
    def _iter_elements(self, root: ET.Element):
        """递归迭代所有元素"""
        yield root
        for child in root:
            yield from self._iter_elements(child)
    
    def parse(self, file_path: str, chunk_size: int = 10000) -> Iterator[List[LogEntry]]:
        chunk: List[LogEntry] = []
        line_num = 0
        
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            for elem in self._iter_elements(root):
                line_num += 1
                entry = self._parse_xml_element(elem, line_num)
                chunk.append(entry)
                
                if len(chunk) >= chunk_size:
                    yield chunk
                    chunk = []
        
        except ET.ParseError as e:
            logger.warning(f"XML解析错误: {e}")
            # 尝试作为文本解析
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line_num, line in enumerate(f, 1):
                    if line.strip().startswith('<'):
                        try:
                            elem = ET.fromstring(line.strip())
                            entry = self._parse_xml_element(elem, line_num)
                            chunk.append(entry)
                        except Exception:
                            continue
                        
                        if len(chunk) >= chunk_size:
                            yield chunk
                            chunk = []
        
        if chunk:
            yield chunk
    
    async def parse_async(self, file_path: str, chunk_size: int = 10000) -> AsyncIterator[List[LogEntry]]:
        # XML解析不需要异步，直接使用同步方法
        for chunk in self.parse(file_path, chunk_size):
            yield chunk


class ExcelFormatParser(FormatParser):
    """Excel格式日志解析器"""
    
    format_type = FormatType.EXCEL
    extensions = [".xlsx", ".xls"]
    
    def detect_format(self, file_path: str) -> bool:
        ext = os.path.splitext(file_path)[1].lower()
        return ext in self.extensions
    
    def _parse_excel_row(self, row: tuple, headers: List[str], line_num: int) -> LogEntry:
        """解析Excel行为LogEntry"""
        data = dict(zip(headers, row))
        
        timestamp = data.get("timestamp") or data.get("time") or data.get("datetime")
        level = (data.get("level") or data.get("severity") or "INFO").upper()
        message = data.get("message") or data.get("msg") or data.get("content") or str(data)
        class_name = data.get("class") or data.get("logger") or ""
        
        return LogEntry(
            timestamp=timestamp,
            level=level,
            message=message,
            class_name=class_name,
            raw_data=data,
            line_number=line_num
        )
    
    def parse(self, file_path: str, chunk_size: int = 10000) -> Iterator[List[LogEntry]]:
        chunk: List[LogEntry] = []
        line_num = 0
        
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                headers = None
                
                for row in sheet.iter_rows(values_only=True):
                    line_num += 1
                    
                    if headers is None:
                        # 第一行作为表头
                        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(row)]
                        continue
                    
                    if not any(row):
                        continue
                    
                    entry = self._parse_excel_row(row, headers, line_num)
                    chunk.append(entry)
                    
                    if len(chunk) >= chunk_size:
                        yield chunk
                        chunk = []
            
            wb.close()
        
        except ImportError:
            logger.error("需要安装 openpyxl: pip install openpyxl")
            raise
        except Exception as e:
            logger.error(f"Excel解析错误: {e}")
            raise
        
        if chunk:
            yield chunk
    
    async def parse_async(self, file_path: str, chunk_size: int = 10000) -> AsyncIterator[List[LogEntry]]:
        # Excel解析使用openpyxl，同步方法已足够高效
        for chunk in self.parse(file_path, chunk_size):
            yield chunk


class PDFFormatParser(FormatParser):
    """PDF格式日志解析器"""
    
    format_type = FormatType.PDF
    extensions = [".pdf"]
    
    # PDF中的常见日志模式
    LOG_PATTERN = re.compile(
        r'(\d{4}-\d{2}-\d{2}[\s\d:,]+)'  # timestamp
        r'\[?\s*(DEBUG|INFO|WARN|WARNING|ERROR|FATAL)\s*\]?'  # level
        r'[\s-]*(.+?)(?=\d{4}-\d{2}-\d{2}[\s\d:,]+\[|\n\n|$)'  # message
    )
    
    def detect_format(self, file_path: str) -> bool:
        ext = os.path.splitext(file_path)[1].lower()
        if ext != ".pdf":
            return False
        try:
            with open(file_path, 'rb') as f:
                header = f.read(5)
                return header.startswith(b'%PDF-')
        except Exception:
            return False
    
    def _extract_text_from_pdf(self, file_path: str) -> List[str]:
        """从PDF提取文本行"""
        lines = []
        try:
            import PyPDF2
            
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    text = page.extract_text()
                    if text:
                        lines.extend(text.split('\n'))
        
        except ImportError:
            logger.error("需要安装 PyPDF2: pip install PyPDF2")
            raise
        except Exception as e:
            logger.error(f"PDF解析错误: {e}")
            raise
        
        return lines
    
    def _parse_pdf_line(self, line: str, line_num: int) -> Optional[LogEntry]:
        """解析PDF文本行为LogEntry"""
        match = self.LOG_PATTERN.search(line)
        if match:
            groups = match.groups()
            return LogEntry(
                timestamp=groups[0].strip() if len(groups) > 0 else None,
                level=groups[1].upper() if len(groups) > 1 else "INFO",
                message=groups[2].strip() if len(groups) > 2 else line,
                line_number=line_num
            )
        elif line.strip():
            return LogEntry(
                message=line.strip(),
                line_number=line_num
            )
        return None
    
    def parse(self, file_path: str, chunk_size: int = 10000) -> Iterator[List[LogEntry]]:
        chunk: List[LogEntry] = []
        line_num = 0
        
        lines = self._extract_text_from_pdf(file_path)
        
        for line in lines:
            line_num += 1
            entry = self._parse_pdf_line(line, line_num)
            if entry:
                chunk.append(entry)
            
            if len(chunk) >= chunk_size:
                yield chunk
                chunk = []
        
        if chunk:
            yield chunk
    
    async def parse_async(self, file_path: str, chunk_size: int = 10000) -> AsyncIterator[List[LogEntry]]:
        for chunk in self.parse(file_path, chunk_size):
            yield chunk


class FormatParserFactory:
    """格式解析器工厂"""
    
    _parsers: List[FormatParser] = [
        LogFormatParser(),
        JSONFormatParser(),
        CSVFormatParser(),
        XMLFormatParser(),
        ExcelFormatParser(),
        PDFFormatParser(),
    ]
    
    @classmethod
    def get_parser(cls, file_path: str) -> FormatParser:
        """根据文件路径获取合适的解析器"""
        ext = os.path.splitext(file_path)[1].lower()
        
        # 首先尝试通过扩展名匹配
        for parser in cls._parsers:
            if ext in parser.extensions:
                if parser.detect_format(file_path):
                    return parser
        
        # 如果扩展名不匹配，尝试自动检测
        for parser in cls._parsers:
            if parser.detect_format(file_path):
                return parser
        
        # 默认使用日志解析器
        return LogFormatParser()
    
    @classmethod
    def detect_format(cls, file_path: str) -> FormatType:
        """自动检测文件格式"""
        for parser in cls._parsers:
            if parser.detect_format(file_path):
                return parser.format_type
        return FormatType.UNKNOWN
    
    @classmethod
    def get_supported_formats(cls) -> List[str]:
        """获取所有支持的格式扩展名"""
        formats = []
        for parser in cls._parsers:
            formats.extend(parser.extensions)
        return formats


def get_universal_parser(file_path: str) -> FormatParser:
    """获取通用解析器（便捷函数）"""
    return FormatParserFactory.get_parser(file_path)


def detect_file_format(file_path: str) -> FormatType:
    """检测文件格式（便捷函数）"""
    return FormatParserFactory.detect_format(file_path)
